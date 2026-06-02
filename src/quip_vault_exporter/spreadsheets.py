"""Spreadsheet export (Phase 3).

Quip spreadsheets export as XLSX via the API. **CSV is NOT a native Quip export** — it is
derived locally from the XLSX with openpyxl. When a spreadsheet can't be represented as
Markdown, we still write a Markdown wrapper note that links to the external files, so the
spreadsheet has a first-class presence in the Obsidian vault.
"""

from __future__ import annotations

import csv
import io
import logging
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .config import Config
from .obsidian import now_iso
from .utils import atomic_write_bytes, atomic_write_text

log = logging.getLogger("quip_vault_exporter.spreadsheets")


@dataclass
class SpreadsheetResult:
    status: str  # "complete" | "failed"
    files: list[str] = field(default_factory=list)
    markdown: str = ""  # wrapper-note content (for checksumming in state)
    error: str | None = None


def xlsx_to_csv(data: bytes) -> str:
    """Render the first worksheet of an XLSX workbook as CSV text."""
    wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    ws = wb.active
    buf = io.StringIO()
    writer = csv.writer(buf)
    if ws is not None:
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if c is None else c for c in row])
    wb.close()
    return buf.getvalue()


def export_spreadsheet(
    *,
    client: Any,
    thread_id: str,
    title: str,
    stem: str,
    folder_dir: Path,
    config: Config,
) -> SpreadsheetResult:
    """Export a spreadsheet thread to XLSX + derived CSV, plus a Markdown wrapper note."""
    files: list[str] = []
    try:
        xlsx = client.export_xlsx(thread_id)
    except Exception as exc:
        return SpreadsheetResult("failed", error=f"xlsx export failed: {exc}")

    atomic_write_bytes(folder_dir / f"{stem}.xlsx", xlsx)
    files.append(f"{stem}.xlsx")

    try:
        csv_text = xlsx_to_csv(xlsx)
        atomic_write_text(folder_dir / f"{stem}.csv", csv_text)
        files.append(f"{stem}.csv")
    except Exception as exc:  # XLSX saved regardless; CSV is best-effort
        log.warning("CSV derivation failed for %s: %s", thread_id, exc)

    wrapper = ""
    if config.export.markdown:
        wrapper = _wrapper_markdown(title, thread_id, files)
        atomic_write_text(folder_dir / f"{stem}.md", wrapper)

    return SpreadsheetResult("complete", files=files, markdown=wrapper)


def _wrapper_markdown(title: str, thread_id: str, files: list[str]) -> str:
    lines = [
        "---",
        "source: quip",
        "type: spreadsheet",
        f'thread_id: "{thread_id}"',
        f'title: "{title}"',
        f"exported_at: {now_iso()}",
        "---",
        "",
        f"# {title}",
        "",
        "This spreadsheet was exported as external files (Quip spreadsheets do not convert",
        "cleanly to Markdown). PDF rendering of spreadsheets caps at ~40,000 cells and",
        "excludes charts.",
        "",
        "Files:",
    ]
    lines += [f"- [[{f}]]" for f in files]
    return "\n".join(lines) + "\n"
